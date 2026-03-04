from typing import Dict, List, Optional, Any
from openpyxl import load_workbook
from app.models.schemas import CustomerData, StationInfo, MeterInfo, SubsystemInfo, ComponentInfo
import logging

logger = logging.getLogger(__name__)


class ExcelReader:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = load_workbook(file_path, data_only=True)
        self.station_sheet = None

    def read_customer_data(self) -> CustomerData:
        self._find_station_sheet()
        if not self.station_sheet:
            raise ValueError("未找到'场站信息'sheet页")

        station_info = self._extract_station_info()
        meter_info = self._extract_meter_info()
        subsystems = self._extract_subsystems()
        component_data = self._extract_component_data(subsystems)

        return CustomerData(
            station_info=station_info,
            meter_info=meter_info,
            subsystems=subsystems,
            component_data=component_data
        )

    def _find_station_sheet(self):
        for sheet_name in self.wb.sheetnames:
            if sheet_name == "场站信息":
                self.station_sheet = self.wb[sheet_name]
                logger.info(f"找到场站信息sheet页: {sheet_name}")
                return
        logger.warning("未找到'场站信息'sheet页")

    def _find_section_start(self, section_name: str) -> Optional[tuple]:
        for row in self.station_sheet.iter_rows(min_row=1):
            for cell in row:
                if cell.value and section_name in str(cell.value):
                    logger.info(f"找到'{section_name}'在第{cell.row}行第{cell.column}列")
                    return (cell.row, cell.column)
        logger.warning(f"未找到'{section_name}'")
        return None

    def _extract_station_info(self) -> StationInfo:
        start_info = self._find_section_start("场站基础信息")
        if not start_info:
            raise ValueError("未找到场站基础信息")

        start_row, start_col = start_info

        data = {}
        for row in self.station_sheet.iter_rows(min_row=start_row + 1, max_col=start_col + 1):
            if not row[start_col - 1].value or str(row[start_col - 1].value).strip() == "":
                break
            key = str(row[start_col - 1].value).strip()
            value = str(row[start_col].value).strip() if row[start_col].value else ""
            data[key] = value
            logger.debug(f"场站信息: {key} = {value}")

        return StationInfo(
            name=data.get("名称", ""),
            short_name=data.get("简称", ""),
            timezone=data.get("时区", ""),
            language=data.get("语言", ""),
            address=data.get("场站地址", ""),
            rated_capacity_mwh=data.get("额定容量MWh", ""),
            rated_power_mw=data.get("额定功率MW", ""),
            longitude=data.get("经度", ""),
            latitude=data.get("纬度", ""),
            station_type=data.get("场站类型", "")
        )

    def _extract_meter_info(self) -> Optional[MeterInfo]:
        start_info = self._find_section_start("关口表信息")
        if not start_info:
            logger.info("未找到关口表信息")
            return None

        start_row, start_col = start_info

        data = {}
        for row in self.station_sheet.iter_rows(min_row=start_row + 1, max_col=start_col + 1):
            if not row[start_col - 1].value or str(row[start_col - 1].value).strip() == "":
                break
            key = str(row[start_col - 1].value).strip()
            value = str(row[start_col].value).strip() if row[start_col].value else ""
            data[key] = value
            logger.debug(f"关口表信息: {key} = {value}")

        if not data:
            return None

        return MeterInfo(
            name=data.get("名称", ""),
            meter_type=data.get("类型", ""),
            rated_capacity=data.get("额定容量", ""),
            rated_power=data.get("额定功率", ""),
            manufacturer=data.get("制造厂家", ""),
            model=data.get("设备型号", ""),
            multiplier=data.get("倍率", ""),
            count=data.get("关口表数量", ""),
        )

    def _extract_subsystems(self) -> List[SubsystemInfo]:
        start_info = self._find_section_start("子系统信息")
        if not start_info:
            logger.info("未找到子系统信息")
            return []

        start_row, start_col = start_info
        subsystems = []

        header_row = start_row + 2
        headers = []
        for cell in self.station_sheet[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        data_start_row = header_row + 1
        for row in self.station_sheet.iter_rows(min_row=data_start_row, max_col=len(headers)):
            if not row[0].value or str(row[0].value).strip() == "":
                break

            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = str(cell.value).strip() if cell.value else ""

            try:
                serial_number = int(row_data.get("序号", 0))
            except (ValueError, TypeError):
                serial_number = len(subsystems) + 1

            subsystem = SubsystemInfo(
                serial_number=serial_number,
                name=row_data.get("储能系统名称", ""),
                manufacturer=row_data.get("制造厂家", ""),
                model=row_data.get("型号", ""),
                rated_capacity=row_data.get("额定容量(kwh)", ""),
                rated_power=row_data.get("额定功率(kw)", ""),
                equipment_structure=row_data.get("设备结构", ""),
                incoming_line_name=row_data.get("所属升压站进线名称", ""),
                transformer_count=row_data.get("箱变数量", ""),
                pcs_count=row_data.get("变流器数量", ""),
                cabin_count=row_data.get("舱数量", ""),
                battery_bank_count=row_data.get("电池组数量", ""),
                battery_cluster_count=row_data.get("电池簇数量", ""),
                energy_meter_count=row_data.get("储能表数量", ""),
                auxiliary_meter_count=row_data.get("辅电表数量", ""),
                # 空调信息：新模板使用“液冷空调结构”和“风冷空调数量”，旧模板为“空调设备结构 / 风冷空调数量 / 液冷空调数量”
                air_conditioner_structure=(
                    row_data.get("液冷空调结构", "")
                    or row_data.get("空调设备结构", "")
                ),
                air_cooler_count=row_data.get("风冷空调数量", ""),
                liquid_cooler_count=row_data.get("液冷空调数量", ""),
                fire_suppression_structure=row_data.get("消防设备结构", ""),
                fire_host_count=row_data.get("主机数量", ""),
                fire_detector_count=row_data.get("探测器数量", ""),
                fire_suppressor_count=row_data.get("抑制机数量", "")
            )
            subsystems.append(subsystem)
            logger.debug(f"子系统: {subsystem.name}, 制造厂家: {subsystem.manufacturer}")

        subsystems.sort(key=lambda x: x.serial_number)
        return subsystems

    def _extract_component_data(self, subsystems: List[SubsystemInfo]) -> Dict[str, Dict[str, ComponentInfo]]:
        """
        根据子系统的“制造厂家+型号”从各个部件信息 sheet 中提取数据。
        新规则：
        - 先从子系统信息中收集所有非空的 制造厂家+型号 组合，作为目标 key
        - 遍历除“场站信息”外的所有 sheet：
          - 在第一行查找 “储能系统型号选择*” 所在列
          - 读取该列下方（从第2行开始）的第一个非空单元格，得到一个 “制造厂家+型号” 字符串
          - 若该字符串在目标 key 集合中，则将此 sheet 解析为该 key 的部件信息
        - 若某个 制造厂家+型号 在所有 sheet 中都未找到对应部件信息，则记录告警日志
        """
        # 收集所有需要查找的 制造厂家+型号 组合
        target_keys: Dict[str, tuple[str, str]] = {}
        for sub in subsystems:
            manufacturer = (sub.manufacturer or "").strip()
            model = (sub.model or "").strip()
            if not manufacturer or not model:
                continue
            key = f"{manufacturer}{model}"
            # 如果存在重复的 key，以第一次出现为准（只影响告警文案，不影响数据）
            target_keys.setdefault(key, (manufacturer, model))

        component_data: Dict[str, Dict[str, ComponentInfo]] = {}

        if not target_keys:
            return component_data

        # 遍历所有 sheet，按“储能系统型号选择*”列的值匹配到 制造厂家+型号
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]

            # 场站信息 sheet 只用于基础数据，不包含部件信息，跳过
            if sheet is self.station_sheet:
                continue

            model_key_in_sheet = self._get_model_key_from_sheet(sheet)
            if not model_key_in_sheet:
                continue

            if model_key_in_sheet not in target_keys:
                continue

            if model_key_in_sheet in component_data:
                logger.warning(
                    f"制造厂家+型号 {model_key_in_sheet} 对应多个部件信息sheet页，"
                    f"已使用首个匹配sheet，忽略sheet: {sheet_name}"
                )
                continue

            component_data[model_key_in_sheet] = self._extract_components_from_sheet(sheet)
            logger.info(f"为制造厂家+型号 {model_key_in_sheet} 提取部件信息自 sheet: {sheet_name}")

        # 对未找到部件信息的 制造厂家+型号 进行告警
        for key, (manufacturer, model) in target_keys.items():
            if key not in component_data:
                logger.warning(
                    f"未找到制造厂家和型号对应的部件信息: 制造厂家={manufacturer}, 型号={model}"
                )

        return component_data

    def _get_model_key_from_sheet(self, sheet) -> Optional[str]:
        """
        从给定 sheet 中，根据“储能系统型号选择*”获取对应的 制造厂家+型号 字符串。
        规则优先级：
        1. 在第1行查找包含“储能系统型号选择”的单元格（兼容是否带 *）
           - 优先取同一行中该单元格右侧的第一个非空单元格作为“制造厂家+型号”
        2. 若同一行右侧没有非空值，则退回到旧逻辑：
           - 从该列下一行开始向下查找第一个非空单元格作为“制造厂家+型号”
        """
        header_cell = None
        for cell in sheet[1]:
            if not cell.value:
                continue
            text = str(cell.value).strip()
            if "储能系统型号选择" in text:
                header_cell = cell
                break

        if not header_cell:
            return None

        # 优先：同一行右侧的第一个非空单元格
        row_idx = header_cell.row
        for col_idx in range(header_cell.column + 1, sheet.max_column + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            key = str(value).strip()
            if key:
                return key

        # 兼容旧结构：向下查找该列第一个非空值
        for row_idx in range(header_cell.row + 1, sheet.max_row + 1):
            value = sheet.cell(row=row_idx, column=header_cell.column).value
            if value is None:
                continue
            key = str(value).strip()
            if key:
                return key

        return None

    def _extract_components_from_sheet(self, sheet) -> Dict[str, ComponentInfo]:
        component_types = [
            "箱变信息", "舱信息", "变流器信息", "电池组信息", "电池簇信息",
            "储能表信息", "风冷空调信息", "液冷空调信息", "消防设备信息"
        ]

        components = {}
        for comp_type in component_types:
            start_info = self._find_section_start_in_sheet(sheet, comp_type)
            if not start_info:
                continue

            start_row, start_col = start_info

            data = {}
            for row in sheet.iter_rows(min_row=start_row + 1, max_col=start_col + 1):
                if not row[start_col - 1].value or str(row[start_col - 1].value).strip() == "":
                    break
                key = str(row[start_col - 1].value).strip()
                value = str(row[start_col].value).strip() if row[start_col].value else ""
                data[key] = value

            box_transformer_type = ""
            cooling_system_type = ""
            pcs_model = ""
            pcs_rated_power = ""
            pcs_manufacturer = ""
            cabin_model = ""
            cabin_manufacturer = ""

            if comp_type == "箱变信息":
                for row in sheet.iter_rows(min_row=1, max_col=10):
                    if row[0].value and "箱变类型" in str(row[0].value):
                        box_transformer_type = str(row[1].value).strip() if row[1].value else ""
                    if row[0].value and "冷却系统类型" in str(row[0].value):
                        cooling_system_type = str(row[1].value).strip() if row[1].value else ""

            if comp_type == "变流器信息":
                for row in sheet.iter_rows(min_row=2, max_row=7):
                    if len(row) > 6 and row[6].value and "制造厂家" in str(row[6].value):
                        pcs_manufacturer = str(row[7].value).strip() if len(row) > 7 and row[7].value else ""
                    if len(row) > 6 and row[6].value and "设备型号" in str(row[6].value):
                        pcs_model = str(row[7].value).strip() if len(row) > 7 and row[7].value else ""
                    if len(row) > 6 and row[6].value and "额定功率" in str(row[6].value):
                        pcs_rated_power = str(row[7].value).strip() if len(row) > 7 and row[7].value else ""

            if comp_type == "舱信息":
                if sheet.max_row >= 4 and sheet[4][3].value and "制造厂家*" in str(sheet[4][3].value):
                    cabin_manufacturer = str(sheet[4][4].value).strip() if sheet[4][4].value else ""
                if sheet.max_row >= 5 and sheet[5][3].value and "设备型号*" in str(sheet[5][3].value):
                    cabin_model = str(sheet[5][4].value).strip() if sheet[5][4].value else ""

            components[comp_type] = ComponentInfo(
                component_type=comp_type,
                data=data,
                box_transformer_type=box_transformer_type,
                cooling_system_type=cooling_system_type,
                pcs_model=pcs_model,
                pcs_rated_power=pcs_rated_power,
                pcs_manufacturer=pcs_manufacturer,
                cabin_model=cabin_model,
                cabin_manufacturer=cabin_manufacturer
            )
            logger.debug(f"提取{comp_type}: {len(data)}个字段")

        return components

    def _find_section_start_in_sheet(self, sheet, section_name: str) -> Optional[tuple]:
        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                if cell.value and section_name in str(cell.value):
                    return (cell.row, cell.column)
        return None

    def get_all_sheet_names(self) -> List[str]:
        return self.wb.sheetnames

    def close(self):
        self.wb.close()
