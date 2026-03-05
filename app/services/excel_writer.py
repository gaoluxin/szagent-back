from typing import Dict, List
from openpyxl import load_workbook
from app.models.schemas import CustomerData
from app.core.config import OUTPUT_DIR
import logging
import re

logger = logging.getLogger(__name__)


class ExcelWriter:
    def __init__(self, template_path: str):
        self.template_path = template_path
        self.wb = load_workbook(template_path)
        self.sheets = {
            "1-场站": None,
            "1.1 物理场站": None,
            "2-储能系统": None,
            "3-箱变": None,
            "4-变流器": None,
            "5-舱": None,
            "6-电池组": None,
            "7-电池簇": None,
            "8-空调": None,
            "9-电表": None,
            "10-消防设备": None,
            "11-其他设备": None
        }
        self._load_sheets()

    def _get_component_key(self, subsystem) -> str:
        """
        根据子系统信息生成用于索引 component_data 的 key：制造厂家+型号。
        需与 ExcelReader._extract_component_data 中的规则保持一致。
        """
        manufacturer = (subsystem.manufacturer or "").strip()
        model = (subsystem.model or "").strip()
        if not manufacturer or not model:
            return ""
        return f"{manufacturer}{model}"

    def _load_sheets(self):
        for sheet_name in self.sheets.keys():
            if sheet_name in self.wb.sheetnames:
                self.sheets[sheet_name] = self.wb[sheet_name]
                logger.info(f"加载sheet页: {sheet_name}")
            else:
                logger.warning(f"模板中未找到sheet页: {sheet_name}")

    def write_customer_data(self, customer_data: CustomerData, output_filename: str) -> str:
        self._write_station_sheet(customer_data)
        self._write_physical_station_sheet(customer_data)
        self._write_energy_storage_system_sheet(customer_data)
        self._write_box_transformer_sheet(customer_data)
        self._write_pcs_sheet(customer_data)
        self._write_cabin_sheet(customer_data)
        self._write_battery_bank_sheet(customer_data)
        self._write_battery_cluster_sheet(customer_data)
        self._write_meter_sheet(customer_data)
        self._write_air_conditioner_sheet(customer_data)
        self._write_fire_suppression_sheet(customer_data)

        output_path = OUTPUT_DIR / output_filename
        self.wb.save(output_path)
        logger.info(f"文件已保存到: {output_path}")
        return str(output_path)

    def _write_station_sheet(self, customer_data: CustomerData):
        sheet = self.sheets.get("1-场站")
        if not sheet:
            logger.warning("未找到1-场站sheet页")
            return

        station = customer_data.station_info
        subsystem_count = len(customer_data.subsystems)

        mapping = {
            "名称*": station.name,
            "时区*": station.timezone,
            "语言*": (station.language or "").strip() or "chn",
            "场站详细地址*": station.address,
            "场站额定容量*": station.rated_capacity_mwh,
            "场站额定功率*": station.rated_power_mw,
            "经度*": station.longitude,
            "纬度*": station.latitude,
            "储能系统数量*": str(subsystem_count),
            "所属物理场站*": f"{station.name}物理场站",
            "场站类型*": station.station_type,
            "Scada别名": "",
            "升压等级": "",
            "电网线路名称": ""
        }

        header_row = 3
        headers = []
        for cell in sheet[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        for col_idx, header in enumerate(headers, start=1):
            if header in mapping:
                sheet.cell(row=4, column=col_idx, value=mapping[header])
                logger.debug(f"写入{header}: {mapping[header]}")

    def _write_physical_station_sheet(self, customer_data: CustomerData):
        sheet = self.sheets.get("1.1 物理场站")
        if not sheet:
            logger.warning("未找到1.1 物理场站sheet页")
            return

        station = customer_data.station_info
        physical_station_name = f"{station.name}物理场站"

        header_row = 3
        headers = []
        for cell in sheet[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        data_row = 4
        for col_idx, header in enumerate(headers, start=1):
            if header == "名称*":
                sheet.cell(row=data_row, column=col_idx, value=physical_station_name)
                logger.debug(f"写入物理场站名称: {physical_station_name}")
            elif header == "场站详细地址*":
                sheet.cell(row=data_row, column=col_idx, value=station.address)
            elif header == "经度*":
                sheet.cell(row=data_row, column=col_idx, value=station.longitude)
            elif header == "纬度*":
                sheet.cell(row=data_row, column=col_idx, value=station.latitude)

    def _write_energy_storage_system_sheet(self, customer_data: CustomerData):
        sheet = self.sheets.get("2-储能系统")
        if not sheet:
            logger.warning("未找到2-储能系统sheet页")
            return

        header_row = 3
        headers = []
        for cell in sheet[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        data_start_row = 4
        for subsystem in customer_data.subsystems:
            pcs_connection_type = self._get_pcs_connection_type(subsystem.equipment_structure)
            system_composition = self._get_system_composition(subsystem)

            mapping = {
                "名称*": subsystem.name,
                "制造厂家*": subsystem.manufacturer,
                "型号*": subsystem.model,
                "额定容量*": subsystem.rated_capacity,
                "额定功率*": subsystem.rated_power,
                "所属升压站进线名称*": subsystem.incoming_line_name,
                "PCS连接形式*": pcs_connection_type,
                "序号*": str(subsystem.serial_number),
                "Scada别名": "",
                "已接入系统构成*": system_composition,
                "模型ID": ""
            }

            for col_idx, header in enumerate(headers, start=1):
                if header in mapping:
                    sheet.cell(row=data_start_row, column=col_idx, value=mapping[header])
                    logger.debug(f"写入{header}: {mapping[header]}")

            data_start_row += 1

    def _get_pcs_connection_type(self, equipment_structure: str) -> str:
        structure_map = {
            "系统模式1": "集中式PCS",
            "系统模式2": "组串式PCS",
            "系统模式3": "分散式"
        }
        return structure_map.get(equipment_structure, "")

    def _get_system_composition(self, subsystem) -> str:
        composition_parts = ["ESSVIEW"]

        if subsystem.battery_cluster_count and subsystem.battery_cluster_count.strip() and subsystem.battery_cluster_count.strip() != "0":
            composition_parts.append("DC")

        if subsystem.pcs_count and subsystem.pcs_count.strip() and subsystem.pcs_count.strip() != "0":
            composition_parts.append("AC")

        has_air_cooler = (
            subsystem.air_cooler_count
            and subsystem.air_cooler_count.strip()
            and subsystem.air_cooler_count.strip() != "0"
        )
        # 新模板中不再提供“液冷空调数量”，改为“液冷空调结构”（空调模式1/2）
        ac_structure = (subsystem.air_conditioner_structure or "").strip()
        has_liquid_cooler = any(
            kw in ac_structure for kw in ["空调模式1", "空调模式2"]
        )
        if has_air_cooler or has_liquid_cooler:
            composition_parts.append("ThermalSystem")

        has_fire_host = subsystem.fire_host_count and subsystem.fire_host_count.strip() and subsystem.fire_host_count.strip() != "0"
        has_fire_detector = subsystem.fire_detector_count and subsystem.fire_detector_count.strip() and subsystem.fire_detector_count.strip() != "0"
        has_fire_suppressor = subsystem.fire_suppressor_count and subsystem.fire_suppressor_count.strip() and subsystem.fire_suppressor_count.strip() != "0"
        if has_fire_host or has_fire_detector or has_fire_suppressor:
            composition_parts.append("FireSuppressionSystem")

        if subsystem.battery_bank_count and subsystem.battery_bank_count.strip() and subsystem.battery_bank_count.strip() != "0":
            composition_parts.append("BatteryBankView")

        return ",".join(composition_parts)

    def _parse_float(self, value: str):
        if value is None:
            return None
        s = str(value).strip()
        if not s:
            return None
        m = re.search(r"[-+]?\d*\.?\d+", s)
        if not m:
            return None
        try:
            return float(m.group())
        except ValueError:
            return None

    def _get_pcs_name(
        self,
        station_short_name: str,
        subsystem,
        pcs_index: int,
        name_example: str | None,
    ) -> str:
        """
        根据默认规则和客户收资表“变流器信息/名称示例*”生成变流器名称。

        默认格式：
            场站简称 + 系统序号 + "#系统-XX变流器"
            例：瑞阳1#系统-01变流器

        若“名称示例*”满足以下条件则覆盖默认规则（优先级按顺序）：
        1. 同时包含场站简称和“号系统” → 场站简称 + 系统序号 + "号系统-XX变流器"
        2. 只包含“#系统”              → 系统序号 + "#系统-XX变流器"
        3. 只包含“号系统”             → 系统序号 + "号系统-XX变流器"
        """
        sn = subsystem.serial_number
        seq = f"{pcs_index:02d}"
        p = (name_example or "").strip()

        # 1) 场站简称 + 系统序号 + "号系统-XX变流器"
        if station_short_name and ("号系统" in p) and (station_short_name in p):
            return f"{station_short_name}{sn}号系统-{seq}变流器"

        if station_short_name and ("#系统" in p) and (station_short_name in p):
            return f"{station_short_name}{sn}#系统-{seq}变流器"

        # 2) 系统序号 + "#系统-XX变流器"
        if "#系统" in p:
            return f"{sn}#系统-{seq}变流器"

        # 3) 系统序号 + "号系统-XX变流器"
        if "号系统" in p:
            return f"{sn}号系统-{seq}变流器"

        # 默认：场站简称 + 系统序号 + "#系统-XX变流器"
        return f"{station_short_name}{sn}#系统-{seq}变流器"

    def _get_cabin_name(
        self,
        station_short_name: str,
        subsystem,
        cabin_index: int,
        name_example: str | None,
    ) -> str:
        """
        根据默认规则和客户收资表“舱信息/名称示例*”生成舱名称。

        默认格式：
            场站简称 + 子系统序号 + "#系统-XX舱"
            例：瑞阳1#系统-01舱

        若“名称示例*”满足以下条件则覆盖默认规则（优先级按顺序）：
        1. 同时包含场站简称和“号系统” → 场站简称 + 子系统序号 + "号系统-XX舱"
        2. 只包含“#系统”              → 子系统序号 + "#系统-XX舱"
        3. 只包含“号系统”             → 子系统序号 + "号系统-XX舱"
        """
        sn = subsystem.serial_number
        seq = f"{cabin_index:02d}"
        p = (name_example or "").strip()

        # 1) 场站简称 + 子系统序号 + "号系统-XX舱"
        if station_short_name and ("号系统" in p) and (station_short_name in p):
            return f"{station_short_name}{sn}号系统-{seq}舱"
        
        if station_short_name and ("#系统" in p) and (station_short_name in p):
            return f"{station_short_name}{sn}#系统-{seq}舱"
        # 2) 子系统序号 + "#系统-XX舱"
        if "#系统" in p:
            return f"{sn}#系统-{seq}舱"

        # 3) 子系统序号 + "号系统-XX舱"
        if "号系统" in p:
            return f"{sn}号系统-{seq}舱"

        # 默认：场站简称 + 子系统序号 + "#系统-XX舱"
        return f"{station_short_name}{sn}#系统-{seq}舱"

    def _get_battery_bank_name(
        self,
        station_short_name: str,
        subsystem,
        bank_index: int,
        name_example: str | None,
    ) -> str:
        """
        根据默认规则和客户收资表“电池组信息/名称示例*”生成电池组名称。

        默认格式：
            场站简称 + 系统序号 + "#系统-XX电池组"
            例：瑞阳1#系统-01电池组

        若“名称示例*”满足以下条件则覆盖默认规则（优先级按顺序）：
        1. 同时包含场站简称和“号系统” → 场站简称 + 系统序号 + "号系统-XX电池组"
        2. 只包含“#系统”              → 系统序号 + "#系统-XX电池组"
        3. 只包含“号系统”             → 系统序号 + "号系统-XX电池组"
        """
        sn = subsystem.serial_number
        seq = f"{bank_index:02d}"
        p = (name_example or "").strip()

        # 1) 场站简称 + 系统序号 + "号系统-XX电池组"
        if station_short_name and ("号系统" in p) and (station_short_name in p):
            return f"{station_short_name}{sn}号系统-{seq}电池组"

        if station_short_name and ("#系统" in p) and (station_short_name in p):
            return f"{station_short_name}{sn}#系统-{seq}电池组"

        # 2) 系统序号 + "#系统-XX电池组"
        if "#系统" in p:
            return f"{sn}#系统-{seq}电池组"

        # 3) 系统序号 + "号系统-XX电池组"
        if "号系统" in p:
            return f"{sn}号系统-{seq}电池组"

        # 默认：场站简称 + 系统序号 + "#系统-XX电池组"
        return f"{station_short_name}{sn}#系统-{seq}电池组"

    def _write_box_transformer_sheet(self, customer_data: CustomerData):
        sheet = self.sheets.get("3-箱变")
        if not sheet:
            logger.warning("未找到3-箱变sheet页")
            return

        header_row = 3
        headers = []
        for cell in sheet[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        data_start_row = 4
        station_short_name = customer_data.station_info.short_name

        for subsystem in customer_data.subsystems:
            try:
                transformer_count = int(subsystem.transformer_count)
            except (ValueError, TypeError):
                transformer_count = 0

            if transformer_count <= 0:
                continue

            # 默认名称格式：场站简称 + 子系统序号 + "#箱变"
            def _build_box_name(pattern: str) -> str:
                """
                根据“名称示例*”字段选中的示例文本生成箱变名称。
                注意：单元格的值是数据验证列表中的实际示例文本（第一项、第二项等），
                而不是“选项1/选项2”这些字样，因此完全按文本内容来判断：
                1. 文本中既包含场站简称又包含“#箱变”   → 场站简称 + 子系统序号 + "#箱变"
                2. 文本中既包含场站简称又包含“号箱变” → 场站简称 + 子系统序号 + "号箱变"
                3. 否则若包含“#箱变”                  → 子系统序号 + "#箱变"
                4. 否则若包含“号箱变”                → 子系统序号 + "号箱变"
                """
                sn = subsystem.serial_number
                p = (pattern or "").strip()

                # 1) 场站简称 + 子系统序号 + "#箱变"
                if station_short_name and ("#箱变" in p) and (station_short_name in p):
                    return f"{station_short_name}{sn}#箱变"

                # 2) 场站简称 + 子系统序号 + "号箱变"
                if station_short_name and ("号箱变" in p) and (station_short_name in p):
                    return f"{station_short_name}{sn}号箱变"

                # 3) 子系统序号 + "#箱变"
                if "#箱变" in p:
                    return f"{sn}#箱变"

                # 4) 子系统序号 + "号箱变"
                if "号箱变" in p:
                    return f"{sn}号箱变"

                # 回退默认：场站简称 + 子系统序号 + "#箱变"
                return f"{station_short_name}{sn}#箱变"

            box_transformer_name = f"{station_short_name}{subsystem.serial_number}#箱变"

            box_transformer_type = ""
            cooling_system_type = ""
            box_transformer_manufacturer = subsystem.manufacturer
            box_transformer_model = subsystem.model

            component_key = self._get_component_key(subsystem)
            if component_key in customer_data.component_data:
                components = customer_data.component_data[component_key]
                if "箱变信息" in components:
                    transformer_info = components["箱变信息"]
                    box_transformer_type = transformer_info.box_transformer_type
                    cooling_system_type = transformer_info.cooling_system_type
                    data = transformer_info.data or {}
                    # 制造厂家/型号优先从客户收资表“箱变信息”里读取（兼容带 * 与不带 *）
                    box_transformer_manufacturer = (
                        data.get("制造厂家*", "").strip()
                        or data.get("制造厂家", "").strip()
                        or box_transformer_manufacturer
                    )
                    box_transformer_model = (
                        data.get("设备型号*", "").strip()
                        or data.get("设备型号", "").strip()
                        or box_transformer_model
                    )
                    # 根据“名称示例*”字段决定箱变名称格式（若存在则覆盖默认规则）
                    name_example = (data.get("名称示例*", "") or "").strip()
                    if name_example:
                        box_transformer_name = _build_box_name(name_example)

            mapping = {
                "名称*": box_transformer_name,
                "制造厂家*": box_transformer_manufacturer,
                "型号*": box_transformer_model,
                "箱变类型*": box_transformer_type,
                "所属系统*": subsystem.name,
                "EnOS箱变类型*": "双绕组",
                "冷却系统类型*": cooling_system_type,
                "序号*": str(subsystem.serial_number),
                "Scada别名": "",
                "模型ID": ""
            }

            for col_idx, header in enumerate(headers, start=1):
                if header in mapping:
                    sheet.cell(row=data_start_row, column=col_idx, value=mapping[header])
                    logger.debug(f"写入{header}: {mapping[header]}")

            data_start_row += 1

    def _write_pcs_sheet(self, customer_data: CustomerData):
        sheet = self.sheets.get("4-变流器")
        if not sheet:
            logger.warning("未找到4-变流器sheet页")
            return

        header_row = 3
        headers = []
        for cell in sheet[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        data_start_row = 4
        station_short_name = customer_data.station_info.short_name

        for subsystem in customer_data.subsystems:
            try:
                pcs_count = int(subsystem.pcs_count)
            except (ValueError, TypeError):
                pcs_count = 0

            if pcs_count <= 0:
                continue

            try:
                battery_bank_count = int(subsystem.battery_bank_count)
            except (ValueError, TypeError):
                battery_bank_count = 1

            pcs_per_group = pcs_count // battery_bank_count if battery_bank_count > 0 else pcs_count

            pcs_model = ""
            pcs_rated_power = ""
            pcs_manufacturer = subsystem.manufacturer
            pcs_name_example = ""

            component_key = self._get_component_key(subsystem)
            if component_key in customer_data.component_data:
                components = customer_data.component_data[component_key]
                if "变流器信息" in components:
                    pcs_info = components["变流器信息"]
                    pcs_model = pcs_info.pcs_model
                    pcs_rated_power = pcs_info.pcs_rated_power
                    if pcs_info.pcs_manufacturer:
                        pcs_manufacturer = pcs_info.pcs_manufacturer
                    # 客户收资表中“名称示例*”字段（若存在）用于决定名称格式
                    pcs_name_example = (
                        (pcs_info.data or {}).get("名称示例*", "") or ""
                    ).strip()

            box_transformer_name = f"{station_short_name}{subsystem.serial_number}#箱变"

            for i in range(pcs_count):
                pcs_number = i + 1
                pcs_name = self._get_pcs_name(
                    station_short_name, subsystem, pcs_number, pcs_name_example
                )

                pcs_group_number = ""
                if "组串式" in subsystem.equipment_structure or "系统模式2" in subsystem.equipment_structure:
                    pcs_group_number = f"#{subsystem.serial_number:03d} PCS组"

                mapping = {
                    "名称*": pcs_name,
                    "制造厂家*": pcs_manufacturer,
                    "设备型号*": pcs_model,
                    "所属箱变*": box_transformer_name,
                    "PCS组编号": pcs_group_number,
                    "额定功率*": pcs_rated_power,
                    "PCS类型*": "交流/直流",
                    "所属系统*": subsystem.name,
                    "序号*": f"{pcs_number:02d}",
                    "Scada别名": "",
                    "模型ID": ""
                }

                for col_idx, header in enumerate(headers, start=1):
                    if header in mapping:
                        sheet.cell(row=data_start_row, column=col_idx, value=mapping[header])
                        logger.debug(f"写入{header}: {mapping[header]}")

                data_start_row += 1

    def _write_cabin_sheet(self, customer_data: CustomerData):
        sheet = self.sheets.get("5-舱")
        if not sheet:
            logger.warning("未找到5-舱sheet页")
            return

        header_row = 3
        headers: List[str] = []
        for cell in sheet[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        logger.info(f"舱表头: {headers}")

        # 数据应紧跟表头行（第3行）开始写入
        data_start_row = 4

        station_short_name = customer_data.station_info.short_name

        for subsystem in customer_data.subsystems:
            try:
                cabin_count = int(subsystem.cabin_count)
            except (ValueError, TypeError):
                cabin_count = 0

            logger.info(f"处理子系统: {subsystem.name}, 舱数量: {cabin_count}")

            component_key = self._get_component_key(subsystem)
            cabin_name_example = ""
            if component_key in customer_data.component_data:
                components = customer_data.component_data[component_key]
                if "舱信息" in components:
                    cabin_info = components["舱信息"]
                    data = cabin_info.data or {}
                    # 制造厂家：优先用客户收资表“舱信息”里的“制造厂家*”
                    cabin_manufacturer = (
                        cabin_info.cabin_manufacturer
                        or data.get("制造厂家*", "").strip()
                        or subsystem.manufacturer
                    )
                    # 型号：优先用客户收资表“舱信息”里的“设备型号*”
                    cabin_model = (
                        cabin_info.cabin_model
                        or data.get("设备型号*", "").strip()
                        or subsystem.model
                    )
                    cabin_name_example = (data.get("名称示例*", "") or "").strip()
                    logger.info(
                        f"找到舱信息: cabin_model={cabin_model}, cabin_manufacturer={cabin_manufacturer}"
                    )
                else:
                    cabin_model = subsystem.model
                    cabin_manufacturer = subsystem.manufacturer
                    logger.info(f"未找到舱信息，使用子系统制造厂家和型号: {cabin_manufacturer}, {cabin_model}")
            else:
                cabin_model = subsystem.model
                cabin_manufacturer = subsystem.manufacturer
                logger.info(f"未找到制造厂家的部件信息，使用子系统制造厂家和型号: {cabin_manufacturer}, {cabin_model}")

            for i in range(cabin_count):
                cabin_number = i + 1
                cabin_name = self._get_cabin_name(
                    station_short_name, subsystem, cabin_number, cabin_name_example
                )

                logger.info(f"写入舱: {cabin_name}")

                mapping = {
                    "名称*": cabin_name,
                    # 按当前模板 5-舱 的固定表头：制造厂家 / 型号 无 *
                    "制造厂家": cabin_manufacturer,
                    "型号": cabin_model,
                    # 所属系统填写对应子系统的储能系统名称
                    "所属系统*": subsystem.name,
                    "Scada别名": ""
                }

                for col_idx, header in enumerate(headers, start=1):
                    if header in mapping:
                        sheet.cell(row=data_start_row, column=col_idx, value=mapping[header])
                        logger.debug(f"写入{header}: {mapping[header]}")

                data_start_row += 1

    def _write_battery_bank_sheet(self, customer_data: CustomerData):
        sheet = self.sheets.get("6-电池组")
        if not sheet:
            logger.warning("未找到6-电池组sheet页")
            return

        header_row = 3
        headers: List[str] = []
        for cell in sheet[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        logger.info(f"电池组表头: {headers}")

        data_start_row = 4
        station_short_name = customer_data.station_info.short_name

        for subsystem in customer_data.subsystems:
            mode = subsystem.equipment_structure or ""

            # 系统模式3（分散式）无电池组设备，不需要填写
            if "系统模式3" in mode:
                logger.info(f"子系统 {subsystem.name} 为系统模式3（分散式），跳过电池组写入")
                continue

            try:
                pcs_count = int(subsystem.pcs_count)
            except (ValueError, TypeError):
                pcs_count = 0

            try:
                battery_bank_count = int(subsystem.battery_bank_count)
            except (ValueError, TypeError):
                battery_bank_count = 0

            if battery_bank_count <= 0:
                logger.info(f"子系统: {subsystem.name} 电池组数量为0，跳过")
                continue

            logger.info(
                f"处理子系统: {subsystem.name}, 模式: {mode}, 变流器数量: {pcs_count}, 电池组数量: {battery_bank_count}"
            )

            battery_manufacturer = subsystem.manufacturer
            battery_model = subsystem.model
            battery_rated_capacity = ""
            battery_name_example = ""

            # 从客户收资表 制造厂家+型号 对应的部件信息sheet 的“电池组信息”中读取制造厂家* / 型号* / 额定容量* / 名称示例*
            component_key = self._get_component_key(subsystem)
            pcs_name_example = ""
            if component_key in customer_data.component_data:
                components = customer_data.component_data[component_key]
                if "电池组信息" in components:
                    battery_info = components["电池组信息"]
                    data = battery_info.data or {}
                    battery_manufacturer = (
                        data.get("制造厂家*", "").strip() or battery_manufacturer
                    )
                    battery_model = data.get("型号*", "").strip() or battery_model
                    battery_name_example = (data.get("名称示例*", "") or "").strip()
                    # 兼容旧表头“额定容量*”/“额定容量(kWh)”和新表头“额定容量(kW)*”
                    battery_rated_capacity = (
                        data.get("额定容量(kW)*", "").strip()
                        or data.get("额定容量*", "").strip()
                        or data.get("额定容量(kWh)*", "").strip()
                        or data.get("额定容量(kWh)", "").strip()
                    )
                if "变流器信息" in components:
                    pcs_info = components["变流器信息"]
                    pcs_name_example = (
                        (pcs_info.data or {}).get("名称示例*", "") or ""
                    ).strip()

            # 生成当前子系统下所有 PCS 名称，后续“所属变流器*”引用
            pcs_names: List[str] = []
            if pcs_count > 0:
                for i in range(pcs_count):
                    pcs_number = i + 1
                    pcs_name = self._get_pcs_name(
                        station_short_name, subsystem, pcs_number, pcs_name_example
                    )
                    pcs_names.append(pcs_name)

            def get_owner_pcs_name(bank_index: int) -> str:
                """根据系统模式和电池组索引，计算所属变流器名称。"""
                if pcs_count <= 0 or not pcs_names:
                    return ""

                # 系统模式2（组串式）：多个变流器对应一个电池组
                if "系统模式2" in mode:
                    if battery_bank_count > 0:
                        pcs_per_bank = max(1, pcs_count // battery_bank_count)
                    else:
                        pcs_per_bank = 1
                    pcs_index = min(bank_index * pcs_per_bank, pcs_count - 1)
                    return pcs_names[pcs_index]

                # 系统模式1（集中式）：电池组和变流器一对一或多对一
                if battery_bank_count >= pcs_count and pcs_count > 0:
                    banks_per_pcs = max(1, battery_bank_count // pcs_count)
                    pcs_index = min(bank_index // banks_per_pcs, pcs_count - 1)
                else:
                    # 电池组数量少于变流器数量时，优先从编号最小的变流器开始依次分配
                    pcs_index = min(bank_index, pcs_count - 1)

                return pcs_names[pcs_index]

            for i in range(battery_bank_count):
                bank_number = i + 1
                bank_name = self._get_battery_bank_name(
                    station_short_name, subsystem, bank_number, battery_name_example
                )
                owner_pcs_name = get_owner_pcs_name(i)

                logger.info(
                    f"写入电池组: {bank_name}, 所属变流器: {owner_pcs_name}, 制造厂家: {battery_manufacturer}, 型号: {battery_model}, 额定容量: {battery_rated_capacity}"
                )

                mapping = {
                    "名称*": bank_name,
                    "制造厂家*": battery_manufacturer,
                    "型号*": battery_model,
                    "所属变流器*": owner_pcs_name,
                    "额定容量*": battery_rated_capacity,
                    "Scada别名": "",
                    "模型ID": "",
                }

                for col_idx, header in enumerate(headers, start=1):
                    if header in mapping:
                        sheet.cell(row=data_start_row, column=col_idx, value=mapping[header])
                        logger.debug(f"写入{header}: {mapping[header]}")

                data_start_row += 1

    def _write_battery_cluster_sheet(self, customer_data: CustomerData):
        sheet = self.sheets.get("7-电池簇")
        if not sheet:
            logger.warning("未找到7-电池簇sheet页")
            return

        header_row = 3
        headers: List[str] = []
        for cell in sheet[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        logger.info(f"电池簇表头: {headers}")

        data_start_row = 4
        station_short_name = customer_data.station_info.short_name

        for subsystem in customer_data.subsystems:
            mode = subsystem.equipment_structure or ""

            try:
                cluster_count = int(subsystem.battery_cluster_count)
            except (ValueError, TypeError):
                cluster_count = 0

            if cluster_count <= 0:
                logger.info(f"子系统: {subsystem.name} 电池簇数量为0，跳过")
                continue

            try:
                battery_bank_count = int(subsystem.battery_bank_count)
            except (ValueError, TypeError):
                battery_bank_count = 0

            try:
                cabin_count = int(subsystem.cabin_count)
            except (ValueError, TypeError):
                cabin_count = 0

            try:
                pcs_count = int(subsystem.pcs_count)
            except (ValueError, TypeError):
                pcs_count = 0

            logger.info(
                f"处理子系统: {subsystem.name}, 模式: {mode}, 电池簇数量: {cluster_count}, 电池组数量: {battery_bank_count}, 舱数量: {cabin_count}, PCS数量: {pcs_count}"
            )

            cluster_manufacturer = subsystem.manufacturer
            cluster_model = subsystem.model
            cluster_rated_capacity = ""
            cell_count = ""
            pack_count = ""
            cell_spec = ""

            # 从客户收资表 制造厂家+型号 对应的部件信息sheet 的“电池簇信息”中读取字段
            component_key = self._get_component_key(subsystem)
            cluster_name_example = ""
            if component_key in customer_data.component_data:
                components = customer_data.component_data[component_key]
                if "电池簇信息" in components:
                    cluster_info = components["电池簇信息"]
                    data = cluster_info.data or {}
                    cluster_manufacturer = (
                        data.get("制造厂家*", "").strip() or cluster_manufacturer
                    )
                    cluster_model = (
                        data.get("设备型号*", "").strip()
                        or data.get("型号*", "").strip()
                        or cluster_model
                    )
                    # 兼容旧表头“额定容量*”/“额定容量(kWh)”和新表头“额定容量(kwh)*”
                    cluster_rated_capacity = (
                        data.get("额定容量(kwh)*", "").strip()
                        or data.get("额定容量(kWh)*", "").strip()
                        or data.get("额定容量*", "").strip()
                        or data.get("额定容量(kwh)", "").strip()
                        or data.get("额定容量(kWh)", "").strip()
                    )
                    cell_count = data.get("包含电芯数量", "").strip()
                    pack_count = data.get("包含电池包数量", "").strip()
                    cell_spec = data.get("电芯规格(Ah)", "").strip()
                    cluster_name_example = (data.get("名称示例*", "") or "").strip()

            # 计算充放电额定功率
            sys_power = self._parse_float(subsystem.rated_power)
            sys_capacity = self._parse_float(subsystem.rated_capacity)
            cluster_capacity_num = self._parse_float(cluster_rated_capacity)

            charge_power = ""
            discharge_power = ""
            if sys_power is not None and sys_capacity not in (None, 0) and cluster_capacity_num is not None:
                ratio = cluster_capacity_num / sys_capacity
                power_value = sys_power * ratio
                power_str = f"{power_value:.4f}".rstrip("0").rstrip(".")
                charge_power = power_str
                discharge_power = power_str

            # 计算充放电额定电压
            cell_spec_num = self._parse_float(cell_spec)
            charge_voltage = ""
            discharge_voltage = ""
            if cluster_capacity_num not in (None, 0) and cell_spec_num not in (None, 0):
                voltage_value = cluster_capacity_num / cell_spec_num
                voltage_str = f"{voltage_value:.4f}".rstrip("0").rstrip(".")
                charge_voltage = voltage_str
                discharge_voltage = voltage_str

            # 电池簇类型 / 是否远景电池簇
            cluster_type = "磷酸铁锂"
            is_yuanjing = "是" if any(
                kw in (cluster_manufacturer or "") for kw in ["远景", "远景能源"]
            ) else "否"

            # 准备所属设备名称列表
            pcs_names: List[str] = []
            component_key = self._get_component_key(subsystem)
            pcs_name_example = ""
            cabin_name_example = ""
            if component_key in customer_data.component_data:
                components = customer_data.component_data[component_key]
                if "变流器信息" in components:
                    pcs_info = components["变流器信息"]
                    pcs_name_example = (
                        (pcs_info.data or {}).get("名称示例*", "") or ""
                    ).strip()
                if "舱信息" in components:
                    cabin_info = components["舱信息"]
                    cabin_name_example = (
                        (cabin_info.data or {}).get("名称示例*", "") or ""
                    ).strip()
            if pcs_count > 0:
                for i in range(pcs_count):
                    pcs_number = i + 1
                    pcs_name = self._get_pcs_name(
                        station_short_name, subsystem, pcs_number, pcs_name_example
                    )
                    pcs_names.append(pcs_name)

            cabin_names: List[str] = []
            if cabin_count > 0:
                for i in range(cabin_count):
                    cabin_number = i + 1
                    cabin_name = self._get_cabin_name(
                        station_short_name, subsystem, cabin_number, cabin_name_example
                    )
                    cabin_names.append(cabin_name)

            bank_names: List[str] = []
            battery_name_example = ""
            component_key = self._get_component_key(subsystem)
            if component_key in customer_data.component_data:
                components = customer_data.component_data[component_key]
                if "电池组信息" in components:
                    battery_info = components["电池组信息"]
                    battery_name_example = (
                        (battery_info.data or {}).get("名称示例*", "") or ""
                    ).strip()
            if battery_bank_count > 0:
                for i in range(battery_bank_count):
                    bank_number = i + 1
                    bank_name = self._get_battery_bank_name(
                        station_short_name, subsystem, bank_number, battery_name_example
                    )
                    bank_names.append(bank_name)

            def get_owner_bank_index(cluster_index: int) -> int:
                """根据电池簇数量和电池组数量的倍数关系，计算所属电池组索引。"""
                if battery_bank_count <= 0:
                    return 0
                if cluster_count >= battery_bank_count:
                    clusters_per_bank = max(1, cluster_count // battery_bank_count)
                    return min(cluster_index // clusters_per_bank, battery_bank_count - 1)
                # 电池簇数量少于电池组数量时，优先从编号最小的电池组开始分配
                return min(cluster_index, battery_bank_count - 1)

            def get_owner_cabin_index(cluster_index: int) -> int:
                """根据电池簇数量和舱数量的倍数关系，计算所属舱索引。"""
                if cabin_count <= 0:
                    return 0
                if cluster_count >= cabin_count:
                    clusters_per_cabin = max(1, cluster_count // cabin_count)
                    return min(cluster_index // clusters_per_cabin, cabin_count - 1)
                # 电池簇数量少于舱数量时，优先从编号最小的舱开始分配
                return min(cluster_index, cabin_count - 1)

            def get_owner_pcs_index_for_mode3(cluster_index: int) -> int:
                """系统模式3（分散式）下，电池簇与变流器一对一。"""
                if pcs_count <= 0:
                    return 0
                return min(cluster_index, pcs_count - 1)

            for i in range(cluster_count):
                # 电池组编号：模式1/2 使用真实电池组编号；模式3 没有电池组，默认为 1
                if "系统模式3" in mode or battery_bank_count <= 0:
                    bank_index = 0
                else:
                    bank_index = get_owner_bank_index(i)

                group_no = bank_index + 1
                # 每个电池组内的电池簇序号（两位），按电池组分别重新编号
                if not hasattr(self, "_battery_cluster_group_seq"):
                    self._battery_cluster_group_seq = {}
                group_key = (subsystem.serial_number, bank_index)
                current_seq = self._battery_cluster_group_seq.get(group_key, 0) + 1
                self._battery_cluster_group_seq[group_key] = current_seq
                seq_code = f"{group_no}{current_seq:02d}"

                # 电池簇名称：默认“场站简称+子系统序号+#系统-XXX电池簇”，可由“电池簇信息/名称示例*”覆盖
                if station_short_name:
                    default_cluster_name = f"{station_short_name}{subsystem.serial_number}#系统-{seq_code}电池簇"
                else:
                    default_cluster_name = f"{subsystem.serial_number}#系统-{seq_code}电池簇"

                p = cluster_name_example
                if station_short_name and ("号系统" in p) and (station_short_name in p):
                    cluster_name = f"{station_short_name}{subsystem.serial_number}号系统-{seq_code}电池簇"
                elif station_short_name and ("#系统" in p) and (station_short_name in p):
                    cluster_name = f"{station_short_name}{subsystem.serial_number}#系统-{seq_code}电池簇"
                elif "#系统" in p:
                    cluster_name = f"{subsystem.serial_number}#系统-{seq_code}电池簇"
                elif "号系统" in p:
                    cluster_name = f"{subsystem.serial_number}号系统-{seq_code}电池簇"
                else:
                    cluster_name = default_cluster_name

                owner_bank_name = ""
                if "系统模式3" not in mode and bank_names:
                    owner_bank_name = bank_names[bank_index]

                owner_pcs_name = ""
                if "系统模式3" in mode and pcs_names:
                    pcs_index = get_owner_pcs_index_for_mode3(i)
                    owner_pcs_name = pcs_names[pcs_index]

                owner_cabin_name = ""
                if cabin_names:
                    cabin_index = get_owner_cabin_index(i)
                    owner_cabin_name = cabin_names[cabin_index]

                logger.info(
                    f"写入电池簇: {cluster_name}, 所属电池组: {owner_bank_name}, 所属PCS: {owner_pcs_name}, 所属舱: {owner_cabin_name}, "
                    f"制造厂家: {cluster_manufacturer}, 型号: {cluster_model}, 额定容量: {cluster_rated_capacity}, 电芯数量: {cell_count}, 电池包数量: {pack_count}, 电芯规格: {cell_spec}"
                )

                mapping = {
                    "名称*": cluster_name,
                    "制造厂家*": cluster_manufacturer,
                    "设备型号*": cluster_model,
                    "额定容量*": cluster_rated_capacity,
                    "充电额定功率*": charge_power,
                    "放电额定功率*": discharge_power,
                    "充电额定电压*": charge_voltage,
                    "放电额定电压*": discharge_voltage,
                    "电池簇类型*": cluster_type,
                    "是否远景电池簇*": is_yuanjing,
                    "所属电池组": owner_bank_name,
                    "所属PCS": owner_pcs_name,
                    "所属舱": owner_cabin_name,
                    "包含电芯数量": cell_count,
                    "序号*": seq_code,
                    "Scada别名": "",
                    "模型ID": "",
                    "包含电池包数量": pack_count,
                    "电芯规格(Ah)": cell_spec,
                }

                for col_idx, header in enumerate(headers, start=1):
                    if header in mapping:
                        sheet.cell(row=data_start_row, column=col_idx, value=mapping[header])
                        logger.debug(f"写入{header}: {mapping[header]}")

                data_start_row += 1

    def _write_air_conditioner_sheet(self, customer_data: CustomerData):
        sheet = self.sheets.get("8-空调")
        if not sheet:
            logger.warning("未找到8-空调sheet页")
            return

        header_row = 3
        headers: List[str] = []
        for cell in sheet[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        logger.info(f"空调表头: {headers}")

        data_start_row = 4
        station_short_name = customer_data.station_info.short_name

        def _get_ac_field(d: Dict[str, str], key_candidates: List[str], fallback_contains: str) -> str:
            for k in key_candidates:
                v = (d.get(k) or "").strip()
                if v:
                    return v
            for k, v in d.items():
                if fallback_contains in (k or "") and (v or "").strip():
                    return (v or "").strip()
            return ""

        for subsystem in customer_data.subsystems:
            # 新模板：子系统“空调信息”提供
            # - 风冷空调数量（空气冷却数量）
            # - 液冷空调结构（空调模式1/2，决定液冷空调层级）
            liquid_structure = (subsystem.air_conditioner_structure or "").strip()
            try:
                wind_count = int(subsystem.air_cooler_count)
            except (ValueError, TypeError):
                wind_count = 0

            try:
                cabin_count = int(subsystem.cabin_count)
            except (ValueError, TypeError):
                cabin_count = 0
            try:
                battery_bank_count = int(subsystem.battery_bank_count)
            except (ValueError, TypeError):
                battery_bank_count = 0
            try:
                cluster_count = int(subsystem.battery_cluster_count)
            except (ValueError, TypeError):
                cluster_count = 0

            if wind_count <= 0 and not liquid_structure:
                logger.info(f"子系统: {subsystem.name} 无空调配置，跳过")
                continue

            # 制造厂家*、设备型号*：
            # - 风冷空调来自“风冷空调信息”
            # - 液冷空调来自“液冷空调信息”
            wind_manufacturer = subsystem.manufacturer
            wind_model = subsystem.model
            wind_name_example = ""
            liquid_manufacturer = subsystem.manufacturer
            liquid_model = subsystem.model
            pack_count_from_cluster = ""

            component_key = self._get_component_key(subsystem)
            cluster_name_example = ""
            if component_key in customer_data.component_data:
                components = customer_data.component_data[component_key]
                if "风冷空调信息" in components:
                    d = components["风冷空调信息"].data or {}
                    v = _get_ac_field(d, ["制造厂家*", "制造厂家"], "制造厂家")
                    if v:
                        wind_manufacturer = v
                    v = _get_ac_field(d, ["设备型号*", "设备型号"], "设备型号")
                    if v:
                        wind_model = v
                    wind_name_example = (d.get("名称示例*", "") or "").strip()
                if "液冷空调信息" in components:
                    d = components["液冷空调信息"].data or {}
                    v = _get_ac_field(d, ["制造厂家*", "制造厂家"], "制造厂家")
                    if v:
                        liquid_manufacturer = v
                    v = _get_ac_field(d, ["设备型号*", "设备型号"], "设备型号")
                    if v:
                        liquid_model = v
                if "电池簇信息" in components:
                    data = components["电池簇信息"].data or {}
                    pack_count_from_cluster = data.get("包含电池包数量", "").strip()
                    cluster_name_example = (data.get("名称示例*", "") or "").strip()

            cabin_names: List[str] = []
            component_key = self._get_component_key(subsystem)
            cabin_name_example = ""
            if component_key in customer_data.component_data:
                components = customer_data.component_data[component_key]
                if "舱信息" in components:
                    cabin_info = components["舱信息"]
                    cabin_name_example = (
                        (cabin_info.data or {}).get("名称示例*", "") or ""
                    ).strip()
            if cabin_count > 0:
                for i in range(cabin_count):
                    cabin_names.append(
                        self._get_cabin_name(
                            station_short_name, subsystem, i + 1, cabin_name_example
                        )
                    )

            bank_names: List[str] = []
            battery_name_example = ""
            component_key = self._get_component_key(subsystem)
            if component_key in customer_data.component_data:
                components = customer_data.component_data[component_key]
                if "电池组信息" in components:
                    battery_info = components["电池组信息"]
                    battery_name_example = (
                        (battery_info.data or {}).get("名称示例*", "") or ""
                    ).strip()
            if battery_bank_count > 0:
                for i in range(battery_bank_count):
                    bank_names.append(
                        self._get_battery_bank_name(
                            station_short_name, subsystem, i + 1, battery_name_example
                        )
                    )

            def get_owner_cabin_index_by_cluster(cluster_index: int) -> int:
                if cabin_count <= 0:
                    return 0
                if cluster_count >= cabin_count:
                    per_cabin = max(1, cluster_count // cabin_count)
                    return min(cluster_index // per_cabin, cabin_count - 1)
                return min(cluster_index, cabin_count - 1)

            def get_owner_cabin_index_by_bank(bank_index: int) -> int:
                if cabin_count <= 0:
                    return 0
                if battery_bank_count >= cabin_count and cabin_count > 0:
                    per_cabin = max(1, battery_bank_count // cabin_count)
                    return min(bank_index // per_cabin, cabin_count - 1)
                return min(bank_index, cabin_count - 1)

            def get_group_no(bank_index: int) -> int:
                # NNN 的首位为电池组编号；分散式无电池组时固定为 1
                if battery_bank_count <= 0:
                    return 1
                return bank_index + 1

            cover_cluster_count = ""
            if cluster_count > 0 and battery_bank_count > 0:
                cover_cluster_count = str(cluster_count // battery_bank_count)

            # 1) 风冷空调：风冷空调数量>0 时，根据风冷空调数量与电池组数量的倍数关系插入记录
            if wind_count > 0:
                thermal_type = "风冷机组"
                thermal_level = "电池组级"

                # 有电池组则按电池组分配；分散式无电池组则视为1组
                effective_bank_count = battery_bank_count if battery_bank_count > 0 else 1

                base_per_bank = wind_count // effective_bank_count
                remainder = wind_count % effective_bank_count
                if remainder != 0:
                    logger.warning(
                        f"子系统: {subsystem.name} 风冷空调数量({wind_count})与电池组数量({effective_bank_count})"
                        f"不是整数倍，将尽量平均分配，前{remainder}个电池组多分配1台。"
                    )

                counts_per_bank: List[int] = []
                for bi in range(effective_bank_count):
                    cnt = base_per_bank + (1 if bi < remainder else 0)
                    counts_per_bank.append(cnt)

                for bank_index, count in enumerate(counts_per_bank):
                    if count <= 0:
                        continue
                    group_no = get_group_no(bank_index)
                    for seq_idx in range(count):
                        local_seq = seq_idx + 1
                        seq_code = f"{group_no}{local_seq:02d}"
                        # 名称*：场站简称+系统序号+"#系统-NNN风冷空调"（按“名称示例*”系统前缀规则覆盖）
                        if station_short_name:
                            default_ac_name = (
                                f"{station_short_name}{subsystem.serial_number}#系统-{seq_code}风冷空调"
                            )
                        else:
                            default_ac_name = (
                                f"{subsystem.serial_number}#系统-{seq_code}风冷空调"
                            )
                        p = wind_name_example
                        if station_short_name and ("号系统" in p) and (
                            station_short_name in p
                        ):
                            ac_name = (
                                f"{station_short_name}{subsystem.serial_number}号系统-{seq_code}风冷空调"
                            )
                        elif station_short_name and ("#系统" in p) and (
                            station_short_name in p
                        ):
                            ac_name = (
                                f"{station_short_name}{subsystem.serial_number}#系统-{seq_code}风冷空调"
                            )
                        elif "#系统" in p:
                            ac_name = (
                                f"{subsystem.serial_number}#系统-{seq_code}风冷空调"
                            )
                        elif "号系统" in p:
                            ac_name = (
                                f"{subsystem.serial_number}号系统-{seq_code}风冷空调"
                            )
                        else:
                            ac_name = default_ac_name
                        owner_cabin = ""
                        if cabin_names:
                            cabin_idx = get_owner_cabin_index_by_bank(
                                min(bank_index, len(cabin_names) - 1)
                            )
                            owner_cabin = cabin_names[cabin_idx]
                        owner_node = (
                            bank_names[bank_index]
                            if bank_names and bank_index < len(bank_names)
                            else ""
                        )

                        mapping = {
                            "名称*": ac_name,
                            "制造厂家*": wind_manufacturer,
                            "设备型号*": wind_model,
                            "热管理机组类型*": thermal_type,
                            "热管理机组层级*": thermal_level,
                            "所属舱*": owner_cabin,
                            "覆盖电池包数量*": "",
                            "覆盖电池簇数量*": cover_cluster_count,
                            "所属上级节点*": owner_node,
                            "序号*": seq_code,
                            "Scada别名": "",
                            "模型ID": "",
                        }
                        for col_idx, header in enumerate(headers, start=1):
                            if header in mapping:
                                sheet.cell(
                                    row=data_start_row, column=col_idx, value=mapping[header]
                                )
                        data_start_row += 1

            # 2) 液冷空调：液冷空调结构=空调模式1/2 时，根据模式填充
            has_liquid_mode1 = "空调模式1" in liquid_structure
            has_liquid_mode2 = "空调模式2" in liquid_structure

            if not has_liquid_mode1 and not has_liquid_mode2:
                if wind_count <= 0:
                    logger.info(f"子系统: {subsystem.name} 未配置液冷空调结构，跳过液冷空调")
                continue

            # 模式2：电池簇级液冷，数量=电池簇数量
            if has_liquid_mode2:
                if cluster_count <= 0:
                    logger.info(f"子系统: {subsystem.name} 液冷空调为电池簇级但电池簇数量为0，跳过")
                else:
                    thermal_type = "液冷机组"
                    thermal_level = "电池簇级"

                    def get_owner_bank_index(cluster_index: int) -> int:
                        if battery_bank_count <= 0:
                            return 0
                        if cluster_count >= battery_bank_count:
                            per_bank = max(1, cluster_count // battery_bank_count)
                            return min(cluster_index // per_bank, battery_bank_count - 1)
                        return min(cluster_index, battery_bank_count - 1)

                    ac_cluster_seq: Dict[tuple, int] = {}
                    for i in range(cluster_count):
                        bank_index = get_owner_bank_index(i)
                        group_no = get_group_no(bank_index)
                        key = (subsystem.serial_number, bank_index)
                        ac_cluster_seq[key] = ac_cluster_seq.get(key, 0) + 1
                        seq_code = f"{group_no}{ac_cluster_seq[key]:02d}"

                        # 与 7-电池簇 一致的电池簇命名规则（受“电池簇信息/名称示例*”影响）
                        if station_short_name:
                            default_cluster_name = (
                                f"{station_short_name}{subsystem.serial_number}#系统-{seq_code}电池簇"
                            )
                        else:
                            default_cluster_name = (
                                f"{subsystem.serial_number}#系统-{seq_code}电池簇"
                            )
                        p = cluster_name_example
                        if station_short_name and ("号系统" in p) and (
                            station_short_name in p
                        ):
                            cluster_name = (
                                f"{station_short_name}{subsystem.serial_number}号系统-{seq_code}电池簇"
                            )
                        elif station_short_name and ("#系统" in p) and (
                            station_short_name in p
                        ):
                            cluster_name = (
                                f"{station_short_name}{subsystem.serial_number}#系统-{seq_code}电池簇"
                            )
                        elif "#系统" in p:
                            cluster_name = (
                                f"{subsystem.serial_number}#系统-{seq_code}电池簇"
                            )
                        elif "号系统" in p:
                            cluster_name = (
                                f"{subsystem.serial_number}号系统-{seq_code}电池簇"
                            )
                        else:
                            cluster_name = default_cluster_name

                        # 液冷空调名称：场站简称+系统序号+"#系统-NNN液冷空调"，同样受系统前缀规则影响
                        if station_short_name:
                            default_ac_name = (
                                f"{station_short_name}{subsystem.serial_number}#系统-{seq_code}液冷空调"
                            )
                        else:
                            default_ac_name = (
                                f"{subsystem.serial_number}#系统-{seq_code}液冷空调"
                            )
                        if station_short_name and ("号系统" in p) and (
                            station_short_name in p
                        ):
                            ac_name = (
                                f"{station_short_name}{subsystem.serial_number}号系统-{seq_code}液冷空调"
                            )
                        elif station_short_name and ("#系统" in p) and (
                            station_short_name in p
                        ):
                            ac_name = (
                                f"{station_short_name}{subsystem.serial_number}#系统-{seq_code}液冷空调"
                            )
                        elif "#系统" in p:
                            ac_name = (
                                f"{subsystem.serial_number}#系统-{seq_code}液冷空调"
                            )
                        elif "号系统" in p:
                            ac_name = (
                                f"{subsystem.serial_number}号系统-{seq_code}液冷空调"
                            )
                        else:
                            ac_name = default_ac_name
                        cabin_index = get_owner_cabin_index_by_cluster(i)
                        owner_cabin = cabin_names[cabin_index] if cabin_names else ""

                        mapping = {
                            "名称*": ac_name,
                            "制造厂家*": liquid_manufacturer,
                            "设备型号*": liquid_model,
                            "热管理机组类型*": thermal_type,
                            "热管理机组层级*": thermal_level,
                            "所属舱*": owner_cabin,
                            "覆盖电池包数量*": pack_count_from_cluster,
                            "覆盖电池簇数量*": "",
                            "所属上级节点*": cluster_name,
                            "序号*": seq_code,
                            "Scada别名": "",
                            "模型ID": "",
                        }
                        for col_idx, header in enumerate(headers, start=1):
                            if header in mapping:
                                sheet.cell(
                                    row=data_start_row, column=col_idx, value=mapping[header]
                                )
                        data_start_row += 1

            # 模式1：电池组级液冷，数量=电池组数量
            if has_liquid_mode1:
                if battery_bank_count <= 0:
                    logger.info(f"子系统: {subsystem.name} 液冷空调为电池组级但电池组数量为0，按1组处理")
                    effective_bank_count = 1
                else:
                    effective_bank_count = battery_bank_count

                thermal_type = "液冷机组"
                thermal_level = "电池组级"

                for bank_index in range(effective_bank_count):
                    group_no = get_group_no(bank_index)
                    # 每个电池组一台液冷空调，组内编号从01开始
                    seq_code = f"{group_no}01"
                    # 电池组级液冷空调名称同样采用系统前缀规则
                    if station_short_name:
                        default_ac_name = (
                            f"{station_short_name}{subsystem.serial_number}#系统-{seq_code}液冷空调"
                        )
                    else:
                        default_ac_name = (
                            f"{subsystem.serial_number}#系统-{seq_code}液冷空调"
                        )
                    p = cluster_name_example
                    if station_short_name and ("号系统" in p) and (
                        station_short_name in p
                    ):
                        ac_name = (
                            f"{station_short_name}{subsystem.serial_number}号系统-{seq_code}液冷空调"
                        )
                    elif station_short_name and ("#系统" in p) and (
                        station_short_name in p
                    ):
                        ac_name = (
                            f"{station_short_name}{subsystem.serial_number}#系统-{seq_code}液冷空调"
                        )
                    elif "#系统" in p:
                        ac_name = f"{subsystem.serial_number}#系统-{seq_code}液冷空调"
                    elif "号系统" in p:
                        ac_name = f"{subsystem.serial_number}号系统-{seq_code}液冷空调"
                    else:
                        ac_name = default_ac_name
                    owner_cabin = ""
                    if cabin_names:
                        cabin_index = get_owner_cabin_index_by_bank(
                            min(bank_index, len(cabin_names) - 1)
                        )
                        owner_cabin = cabin_names[cabin_index]
                    owner_node = (
                        bank_names[bank_index]
                        if bank_names and bank_index < len(bank_names)
                        else ""
                    )

                    mapping = {
                        "名称*": ac_name,
                        "制造厂家*": liquid_manufacturer,
                        "设备型号*": liquid_model,
                        "热管理机组类型*": thermal_type,
                        "热管理机组层级*": thermal_level,
                        "所属舱*": owner_cabin,
                        "覆盖电池包数量*": "",
                        "覆盖电池簇数量*": cover_cluster_count,
                        "所属上级节点*": owner_node,
                        "序号*": seq_code,
                        "Scada别名": "",
                        "模型ID": "",
                    }
                    for col_idx, header in enumerate(headers, start=1):
                        if header in mapping:
                            sheet.cell(
                                row=data_start_row, column=col_idx, value=mapping[header]
                            )
                    data_start_row += 1

    def _write_fire_suppression_sheet(self, customer_data: CustomerData):
        sheet = self.sheets.get("10-消防设备")
        if not sheet:
            logger.warning("未找到10-消防设备sheet页")
            return

        header_row = 3
        headers: List[str] = []
        for cell in sheet[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        logger.info(f"消防设备表头: {headers}")

        data_start_row = 4
        station_short_name = customer_data.station_info.short_name

        for subsystem in customer_data.subsystems:
            fire_struct = (subsystem.fire_suppression_structure or "").strip()
            # 只关心是否大于0，具体数量由电池组/电池簇个数决定
            try:
                has_host = int(subsystem.fire_host_count) > 0
            except (ValueError, TypeError):
                has_host = False
            try:
                has_detector = int(subsystem.fire_detector_count) > 0
            except (ValueError, TypeError):
                has_detector = False
            try:
                has_suppressor = int(subsystem.fire_suppressor_count) > 0
            except (ValueError, TypeError):
                has_suppressor = False

            if not (has_host or has_detector or has_suppressor):
                logger.info(f"子系统: {subsystem.name} 无消防设备，跳过")
                continue

            try:
                battery_bank_count = int(subsystem.battery_bank_count)
            except (ValueError, TypeError):
                battery_bank_count = 0
            try:
                cluster_count = int(subsystem.battery_cluster_count)
            except (ValueError, TypeError):
                cluster_count = 0
            mode = subsystem.equipment_structure or ""
            is_dispersed = "系统模式3" in mode

            # 消防设备层级*：结构为1/3时为电池组级，否则为电池簇级
            is_group_level = any(
                x in fire_struct for x in ["1", "3", "消防模式1", "消防模式3"]
            ) and all(x not in fire_struct for x in ["消防模式2", "消防结构2", "2"])
            fire_level = "电池组级" if is_group_level else "电池簇级"
            # 消防设备结构为“消防结构2”或“消防模式2”时，按电池簇级所属上级（电池簇）处理
            is_mode2 = (
                "消防模式2" in fire_struct
                or "消防结构2" in fire_struct
                or fire_struct == "2"
            )
            is_mixed = "混合" in fire_struct

            # 从客户收资表 消防设备信息 读取
            fire_manufacturer = subsystem.manufacturer
            fire_model = subsystem.model
            fire_name_example = ""
            has_pack_detector = ""
            has_cluster_detector = ""
            pack_detector_count_val = ""
            component_key = self._get_component_key(subsystem)
            if component_key in customer_data.component_data:
                comps = customer_data.component_data[component_key]
                if "消防设备信息" in comps:
                    d = (comps["消防设备信息"].data or {})
                    fire_manufacturer = (d.get("制造厂家*") or d.get("制造厂家") or "").strip() or fire_manufacturer
                    fire_model = (d.get("设备型号*") or d.get("设备型号") or "").strip() or fire_model
                    fire_name_example = (d.get("名称示例*", "") or "").strip()
                    has_pack_detector = (d.get("是否包含包级探测器") or "").strip()
                    has_cluster_detector = (d.get("是否包含簇级探测器") or "").strip()
                if "电池簇信息" in comps and (has_pack_detector or "").strip() == "1-yes":
                    pack_detector_count_val = (
                        (comps["电池簇信息"].data or {}).get("包含电池包数量", "")
                    ).strip()

            bank_names: List[str] = []
            battery_name_example = ""
            component_key = self._get_component_key(subsystem)
            if component_key in customer_data.component_data:
                comps = customer_data.component_data[component_key]
                if "电池组信息" in comps:
                    battery_info = comps["电池组信息"]
                    battery_name_example = (
                        (battery_info.data or {}).get("名称示例*", "") or ""
                    ).strip()
            if battery_bank_count > 0:
                for i in range(battery_bank_count):
                    bank_names.append(
                        self._get_battery_bank_name(
                            station_short_name, subsystem, i + 1, battery_name_example
                        )
                    )

            # 电池簇名称与序号（与7-电池簇规则一致，按电池组内从01递增）
            cluster_names: List[str] = []
            cluster_seq_codes: List[str] = []
            if cluster_count > 0:
                def _get_bank_index(cluster_index: int) -> int:
                    if battery_bank_count <= 0:
                        return 0
                    if cluster_count >= battery_bank_count:
                        per = max(1, cluster_count // battery_bank_count)
                        return min(cluster_index // per, battery_bank_count - 1)
                    return min(cluster_index, battery_bank_count - 1)

                seq_map: Dict[tuple, int] = {}
                for i in range(cluster_count):
                    bi = _get_bank_index(i)
                    group_no = bi + 1
                    key = (subsystem.serial_number, bi)
                    seq_map[key] = seq_map.get(key, 0) + 1
                    code = f"{group_no}{seq_map[key]:02d}"
                    cluster_seq_codes.append(code)
                    cluster_names.append(
                        f"{station_short_name}{subsystem.serial_number}#-{code}电池簇"
                    )

            def write_fire_row(
                fire_type: str,
                nnn: str,
                owner_name: str,
            ) -> None:
                nonlocal data_start_row
                # MM：消防控制→控制，消防探测→探测，消防抑制→抑制
                mm = {"消防控制": "控制", "消防探测": "探测", "消防抑制": "抑制"}.get(fire_type, "")

                # 默认名称：场站简称 + 系统序号 + "#系统-NNNMM消防设备"
                if station_short_name:
                    default_name = f"{station_short_name}{subsystem.serial_number}#系统-{nnn}{mm}消防设备"
                else:
                    default_name = f"{subsystem.serial_number}#系统-{nnn}{mm}消防设备"

                p = fire_name_example
                if station_short_name and ("号系统" in p) and (station_short_name in p):
                    name = f"{station_short_name}{subsystem.serial_number}号系统-{nnn}{mm}消防设备"
                elif station_short_name and ("#系统" in p) and (station_short_name in p):
                    name = f"{station_short_name}{subsystem.serial_number}#系统-{nnn}{mm}消防设备"
                elif "#系统" in p:
                    name = f"{subsystem.serial_number}#系统-{nnn}{mm}消防设备"
                elif "号系统" in p:
                    name = f"{subsystem.serial_number}号系统-{nnn}{mm}消防设备"
                else:
                    name = default_name

                mapping = {
                    "名称*": name,
                    "制造厂家*": fire_manufacturer,
                    "设备型号*": fire_model,
                    "消防设备类型*": fire_type,
                    "消防设备层级*": fire_level,
                    "是否包含包级探测器": has_pack_detector,
                    "是否包含簇级探测器": has_cluster_detector,
                    "包级探测器数量": pack_detector_count_val,
                    "所属上级设备*": owner_name,
                    "Scada别名": "",
                    "模型ID": "",
                }
                for col_idx, header in enumerate(headers, start=1):
                    if header in mapping:
                        sheet.cell(row=data_start_row, column=col_idx, value=mapping[header])
                data_start_row += 1

            # 消防模式2：数量=电池簇个数，NNN=电池簇编号(3位)，所属上级=电池簇
            if is_mode2:
                n_cluster = len(cluster_seq_codes)
                if has_host and n_cluster:
                    for i in range(n_cluster):
                        write_fire_row("消防控制", cluster_seq_codes[i], cluster_names[i])
                if has_detector and n_cluster:
                    for i in range(n_cluster):
                        write_fire_row("消防探测", cluster_seq_codes[i], cluster_names[i])
                if has_suppressor and n_cluster:
                    for i in range(n_cluster):
                        write_fire_row("消防抑制", cluster_seq_codes[i], cluster_names[i])
                continue

            # 电池组级：数量=电池组个数，NNN=01,02...，所属上级=电池组
            if is_group_level:
                if has_host and bank_names:
                    for i in range(battery_bank_count):
                        write_fire_row("消防控制", f"{i + 1:02d}", bank_names[i])
                if has_detector and bank_names:
                    for i in range(battery_bank_count):
                        write_fire_row("消防探测", f"{i + 1:02d}", bank_names[i])
                if has_suppressor and bank_names:
                    for i in range(battery_bank_count):
                        write_fire_row("消防抑制", f"{i + 1:02d}", bank_names[i])
                continue

            # 电池簇级（非分散式）或 混合：主机/抑制机数量=电池组个数，探测器数量=电池簇个数
            n_cluster = len(cluster_seq_codes)
            if has_host and bank_names:
                for i in range(battery_bank_count):
                    write_fire_row("消防控制", f"{i + 1:02d}", bank_names[i])
            if has_detector and n_cluster:
                for i in range(n_cluster):
                    write_fire_row("消防探测", cluster_seq_codes[i], cluster_names[i])
            if has_suppressor and bank_names:
                for i in range(battery_bank_count):
                    write_fire_row("消防抑制", f"{i + 1:02d}", bank_names[i])

    def _write_meter_sheet(self, customer_data: CustomerData):
        sheet = self.sheets.get("9-电表")
        if not sheet:
            logger.warning("未找到9-电表sheet页")
            return

        header_row = 3
        headers: List[str] = []
        for cell in sheet[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())

        logger.info(f"电表表头: {headers}")

        data_start_row = 4
        station = customer_data.station_info
        station_short_name = station.short_name

        # 关口表（场站级）
        meter_info = customer_data.meter_info
        if meter_info:
            try:
                gate_count = int(meter_info.count) if meter_info.count else 1
            except (ValueError, TypeError):
                gate_count = 1

            try:
                rated_power_mw = float(str(station.rated_power_mw).strip() or "0")
            except ValueError:
                rated_power_mw = 0.0

            # 斜率*: 按额定功率MW / 2000 计算（若无法解析则留空）
            gate_slope = ""
            if rated_power_mw:
                slope_val = rated_power_mw / 2000.0
                gate_slope = f"{slope_val:.6f}".rstrip("0").rstrip(".")

            for i in range(gate_count):
                idx = i + 1
                name = f"{station_short_name}{idx:02d}关口表"
                mapping = {
                    "名称*": name,
                    "制造厂家*": meter_info.manufacturer,
                    "设备型号*": meter_info.model,
                    "电表所属层级*": "场站级",
                    "类型*": "关口表",
                    "倍率*": meter_info.multiplier,
                    "接入模式*": "正接",
                    "所属系统": "",
                    "斜率*": gate_slope,
                    "序号*": f"{idx:02d}",
                    "Scada别名": "",
                    "模型ID": "",
                }

                for col_idx, header in enumerate(headers, start=1):
                    if header in mapping:
                        sheet.cell(row=data_start_row, column=col_idx, value=mapping[header])
                data_start_row += 1

        # 储能表（系统级）
        for subsystem in customer_data.subsystems:
            try:
                energy_meter_count = int(subsystem.energy_meter_count)
            except (ValueError, TypeError):
                energy_meter_count = 0

            if energy_meter_count <= 0:
                continue

            meter_manufacturer = subsystem.manufacturer
            meter_model = subsystem.model
            meter_multiplier = ""
            meter_name_example = ""

            component_key = self._get_component_key(subsystem)
            if component_key in customer_data.component_data:
                comps = customer_data.component_data[component_key]
                if "储能表信息" in comps:
                    d = comps["储能表信息"].data or {}
                    meter_manufacturer = (
                        (d.get("制造厂家*") or d.get("制造厂家") or meter_manufacturer)
                    ).strip()
                    meter_model = (
                        (d.get("设备型号*") or d.get("设备型号") or meter_model)
                    ).strip()
                    meter_multiplier = (d.get("倍率*") or d.get("倍率") or "").strip()
                    meter_name_example = (d.get("名称示例*", "") or "").strip()

            # 斜率*: 子系统额定功率(kw) / 2（若无法解析则留空）
            try:
                rated_power_kw = float(str(subsystem.rated_power).strip() or "0")
            except ValueError:
                rated_power_kw = 0.0

            storage_slope = ""
            if rated_power_kw:
                slope_val = rated_power_kw / 2.0
                storage_slope = f"{slope_val:.6f}".rstrip("0").rstrip(".")

            for i in range(energy_meter_count):
                idx = i + 1
                seq = f"{idx:02d}"
                
                # 默认：场站简称 + 子系统序号 + "#系统-XX储能表"
                if station_short_name:
                    default_name = (
                        f"{station_short_name}{subsystem.serial_number}#系统-{seq}储能表"
                    )
                else:
                    default_name = f"{subsystem.serial_number}#系统-{seq}储能表"

                p = meter_name_example
                if station_short_name and ("号系统" in p) and (
                    station_short_name in p
                ):
                    name = (
                        f"{station_short_name}{subsystem.serial_number}号系统-{seq}储能表"
                    )
                elif station_short_name and ("#系统" in p) and (
                    station_short_name in p
                ):
                    name = (
                        f"{station_short_name}{subsystem.serial_number}#系统-{seq}储能表"
                    )
                elif "#系统" in p:
                    name = f"{subsystem.serial_number}#系统-{seq}储能表"
                elif "号系统" in p:
                    name = f"{subsystem.serial_number}号系统-{seq}储能表"
                else:
                    name = default_name
                mapping = {
                    "名称*": name,
                    "制造厂家*": meter_manufacturer,
                    "设备型号*": meter_model,
                    "电表所属层级*": "系统级",
                    "类型*": "储能表",
                    "倍率*": meter_multiplier,
                    "接入模式*": "正接",
                    "所属系统": subsystem.name,
                    "斜率*": storage_slope,
                    "序号*": f"{idx:02d}",
                    "Scada别名": "",
                    "模型ID": "",
                }

                for col_idx, header in enumerate(headers, start=1):
                    if header in mapping:
                        sheet.cell(row=data_start_row, column=col_idx, value=mapping[header])
                data_start_row += 1

    def close(self):
        self.wb.close()
