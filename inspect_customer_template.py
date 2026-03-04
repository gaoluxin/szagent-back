from openpyxl import load_workbook


def main():
    wb = load_workbook("public/储能电站收资表-客户版.xlsx", data_only=True)
    try:
        print("=== 所有 sheet 名称 ===")
        for name in wb.sheetnames:
            print(f"- {name}")
        print()

        for name in wb.sheetnames:
            sheet = wb[name]
            print(f"=== sheet: {name} ===")
            # 打印第1行前 15 列内容，查看是否有“储能系统型号选择*”
            row = sheet[1]
            values = []
            for idx, cell in enumerate(row[:15], start=1):
                if cell.value is not None:
                    values.append(f"列{idx}: {cell.value}")
            if values:
                print("第1行非空单元格：")
                for v in values:
                    print("  ", v)
            else:
                print("第1行全部为空")

            # 再尝试在整表中找包含“储能系统型号选择”的单元格，展示其坐标和下面几行的值
            found = False
            for row_cells in sheet.iter_rows(min_row=1, max_row=5):
                for cell in row_cells:
                    if cell.value and "储能系统型号选择" in str(cell.value):
                        found = True
                        col = cell.column
                        print(f"找到 '储能系统型号选择' 在行{cell.row}, 列{col}")
                        for r in range(cell.row + 1, min(cell.row + 6, sheet.max_row + 1)):
                            v = sheet.cell(row=r, column=col).value
                            print(f"  ↓ 第{r}行, 列{col}: {v}")
            if not found:
                print("未在前5行找到包含 '储能系统型号选择' 的单元格")
            print()
    finally:
        wb.close()


if __name__ == "__main__":
    main()

